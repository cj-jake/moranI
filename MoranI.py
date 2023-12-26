import openpyxl
import os
from openpyxl.styles import PatternFill
from KnnImplementSpatialWeightMatrix import *
from calculateMoranI import *


def knnMoran():
    choice_data_column_list_in = []
    # 读入文件
    path = r"/"
    os.chdir(path)  # 修改工作路径
    wb = openpyxl.load_workbook('示例数据.xlsx')
    ws_list = wb.sheetnames
    yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')  # 黄
    green_fill = PatternFill(patternType='solid', fgColor='90EE90')  # 淡绿色
    # 用户选择需使用的工作表
    print(ws_list)
    # choice_sheet = input("请输入使用的表:")
    try:
        ws1 = wb['Sheet1']
        print('当前处理表为', ws1)
        # 获取当前表的实际最大行数和列数
        max_rows, max_columns = 0, 0
        for i in range(1, ws1.max_row + 1):
            if type(ws1.cell(i, 1).value) is int or type(ws1.cell(i, 1).value) is str or type(
                    ws1.cell(i, 1).value) is float:
                max_rows += 1
            else:
                break
        for i in range(1, ws1.max_column + 1):
            if type(ws1.cell(1, i).value) is int or type(ws1.cell(1, i).value) is str or type(
                    ws1.cell(1, i).value) is float:
                max_columns += 1
            else:
                break
        # 用户选择需处理数据
        print("\n表中数据有：")
        for i in range(1, max_columns + 1):
            print(f"{i}:{ws1.cell(1, i).value}\t", end='')
        # choice_data_column = input("\n请输入属性数据序号：")
        # choice_weight_column = input("请输入权重数据序号：")
        # choice_data_column_list = choice_data_column.split(" ")
        # choice_weight_column_list = choice_data_column.split(" ")
        choice_data_column_list_in = [6]
        choice_weight_column_list_in = [2, 3]
        f = open(r"/\空间权重矩阵.txt", "w+")
        data = []
        weight = []
        weight_column_list = []
        # 执行数据处理操作
        for choice_weight_column in choice_weight_column_list_in:
            weight_column_list.append(int(choice_weight_column))
        # print(weight_column_list)
        # 读取Excel表中权重数据
        n = len(weight_column_list)
        for i in range(2, max_rows + 1):
            weight_tem = []
            for j in range(n):
                weight_tem.append(ws1.cell(i, weight_column_list[j]).value)
            weight.append(weight_tem)
        for choice_data_column in choice_data_column_list_in:
            choice_data_column = int(choice_data_column)
            # 读取Excel表中属性数据
            for i in range(2, max_rows + 1):
                data.append(ws1.cell(i, choice_data_column).value)
            try:
                # 计算空间权重矩阵
                w = knn(weight, method=n)
                # 将空间权重矩阵写入txt文件
                f.write(f"{ws1.cell(1, choice_data_column).value}的空间权重矩阵\n")
                for i in range(max_rows - 1):
                    f.write("[")
                    for j in range(max_rows - 1):
                        f.write(str(w[i][j]) + " ")
                    f.write("]\n")
                print(f"\n输入属性数据为：{ws1.cell(1, choice_data_column).value}")
                for weight_column in weight_column_list:
                    print(f"输入权重数据为：{ws1.cell(1, weight_column).value}")
                # 计算莫兰指数
                data = np.array(data)
                data = data / data.sum(axis=0)  # 数据归一化
                moranI = (MoranI(w, data))
                print(f"{ws1.cell(1, choice_data_column).value}的{moranI['I']['desc']}为", moranI["I"]["value"], "\n")
                # 将计算出的莫兰指数写入表中
                ws1.cell(1,
                         max_columns + choice_data_column).value = f"{ws1.cell(1, choice_data_column).value}的{moranI['Ii']['desc']}"
                for i in range(2, max_rows + 1):
                    ws1.cell(i, max_columns + choice_data_column).value = moranI['Ii']['value'][i - 2]
                    if abs(moranI['Ii']['value'][i - 2]) > 1:
                        for j in [1, 4, max_columns+choice_data_column]:
                            ws1.cell(i, j).fill = yellow_fill
                        ws1.cell(i, 5).fill = green_fill
                ws1.cell(max_rows + 1, 1).value = moranI["I"]["desc"]
                ws1.cell(max_rows + 1, max_columns + choice_data_column).value = moranI["I"]['value']
            except ValueError as e:
                print("请检查您的数据！")
                print(e)
    except KeyError as e:
        print("请输入以上列表中指定名称！")
        print(e)
    finally:
        wb.save(r'E:\code\python\others\wrBook\KNN-MoranI\Knn计算结果.xlsx')


# 函数主体
if __name__ == '__main__':
    # data = [1, 6, 5, 7, 2, 1, 3]
    # w = knn(data, method=2)
    # moranI = MoranI(w, data)
    # print(w)
    # print(f"{moranI['I']['desc']}为", moranI["I"]["value"], "\n")
    # print(f"{moranI['Ii']['desc']}为", moranI["Ii"]["value"], "\n")
    knnMoran()